from openpyxl import load_workbook
from salary import cheek_salary
from leave import check_leave
from report import create_report

wb = load_workbook("employee_data.xlsx")    
sheet = wb.active
final_data = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    name = row[0]
    salary = int(row[1])
    leave = int(row[2])
    salary_status = cheek_salary(salary)
    leave_status = check_leave(leave)
    final_data.append({"name": name, "salary_status": salary_status, "leave_status": leave_status})
create_report(final_data)