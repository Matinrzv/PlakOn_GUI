import sqlite3
# conn = sqlite3.connect("employees.db")
# print("Database created")
# conn.close()

# conn = sqlite3.connect("employees.db")
# cursor = conn.cursor()
# cursor.execute("""CREATE TABLE employees (
#                name TEXT,
#                salary INTEGER,
#                leaves INTEGER
#                )""")
# conn.commit()
# conn.close()

# conn = sqlite3.connect("employees.db")
# cursor = conn.cursor()
# cursor.execute("INSERT INTO employees VALUES ('Alice', 5000, 2)")
# cursor.execute("INSERT INTO employees VALUES ('Bob', 6000, 0)")
# cursor.execute("INSERT INTO employees VALUES ('Charlie', 5500, 5)")
# cursor.execute("INSERT INTO employees VALUES ('David', 4500, 1)")
# conn.commit()
# conn.close()

# conn = sqlite3.connect("employees.db")
# cursor = conn.cursor()
# cursor.execute("SELECT * FROM employees")
# rows = cursor.fetchall()
# for row in rows:
#     print(row)
# conn.close()

# conn = sqlite3.connect("employees.db")
# cursor = conn.cursor()
# cursor.execute("SELECT * FROM employees")
# rows = cursor.fetchall()
# for row in rows:
#     name = row[0]
#     salary = row[1]
#     leaves = row[2]
#     if salary > 5000:
#         salary_status = "High"
#     else:
#         salary_status = "Normal"
#     if leaves > 3:
#         leave_status = "Rejected"
#     else:
#         leave_status = "Approved"
#     print(f"{name}: Salary Status: {salary_status}, Leave Status: {leave_status}")

# from openpyxl import load_workbook
# import sqlite3
# wb = load_workbook("./employee_system/employee_data.xlsx")
# sheet = wb.active
# conn = sqlite3.connect("./employee_system/employees.db")
# cursor = conn.cursor()
# for row in sheet.iter_rows(min_row=2, values_only=True):
#     name = row[0]
#     salary = row[1]
#     leave = row[2]
#     cursor.execute("INSERT INTO employees (name, salary, leaves) VALUES (?, ?, ?)", (name, salary, leave))
# conn.commit()
# conn.close()

import sqlite3
from openpyxl import Workbook
conn = sqlite3.connect("./employee_system/employees.db")
cursor = conn.cursor()
cursor.execute("SELECT * FROM employees")
rows = cursor.fetchall()
wb = Workbook()
sheet = wb.active
sheet.append(["Name", "Salary Status", "Leave Status"])
for row in rows:
    name = row[0]
    salary = row[1]
    leave = row[2]
    salary_status = "High" if salary > 5000 else "Normal"
    leave_status = "Rejected" if leave > 3 else "Approved"
    sheet.append([name, salary_status, leave_status])
wb.save("./employee_system/employee_report.xlsx")
conn.close()