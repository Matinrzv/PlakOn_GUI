# import csv
# with open("leaves.csv", "r") as file , open("report.csv", "w") as report:
#     reader = csv.DictReader(file)
#     fieldnames = ["name","status"]
#     writer = csv.DictWriter(report, fieldnames=fieldnames)
#     for row in reader:
#         name = row["name"]
#         leave = int(row["leave"])
#         status = "Approved" if leave > 0 else "Rejected"
#         writer.writerow({"name": name, "status": status})
# from openpyxl import Workbook
# wb = Workbook()
# sheet = wb.active
# sheet["A1"] = "Name"
# sheet["B1"] = "Status"
# sheet["A2"] = "Alice"
# sheet["B2"] = "2"
# sheet["A3"] = "Bob"
# sheet["B3"] = "0"
# sheet["A4"] = "Charlie"
# sheet["B4"] = "5"
# sheet["A5"] = "David"
# sheet["B5"] = "1"
# wb.save("leaves.xlsx")
# from openpyxl import load_workbook
# wb = load_workbook("leaves.xlsx")
# sheet = wb.active
# for row in sheet.iter_rows(min_row=2, values_only=True):
#     name = row[0]
#     leave = int(row[1])
#     print(f"{name}: {leave}")
#     if leave > 3:
#         print(f"{name} has too many leaves. Status: Rejected")
#     else:
#         print(f"{name} has enough leaves. Status: Approved")
from openpyxl import Workbook,load_workbook
wb = load_workbook("leaves.xlsx")
sheet = wb.active
report = Workbook()
report_sheet = report.active
for row in sheet.iter_rows(min_row=2, values_only=True):
    name = row[0]
    leave = int(row[1])
    status = "Approved" if leave <= 3 else "Rejected"
    report_sheet.append([name, status])
report.save("report.xlsx")